import os
from typing import List

import openpyxl as xl
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side



def xlsx_file_read(path):
    # XLSX-file read-write
    attendance_data = {}
    xlsx_file_path = f'{path}/attendance.xlsx'
    if os.path.isfile(xlsx_file_path):
        # Read XLSX-file
        work_book = xl.load_workbook(xlsx_file_path)
        work_sheet = work_book.active
        # Read names
        names = []
        for col in work_sheet.iter_cols(min_col=1, max_col=1, min_row=2,
                                        max_row=work_sheet.max_row):
            for cell in col:
                names.append(cell.value)
        print('names', names)
        # Read dates of attendance
        attendance_dates = []
        for row in work_sheet.iter_rows(min_row=1, max_row=1, min_col=2,
                                        max_col=work_sheet.max_column):
            for cell in row:
                attendance_dates.append(cell.value)
        print('attendance_dates', attendance_dates)
        # Creating a Dictionary from an Excel File

        for att_name, row in zip(names, work_sheet.iter_rows(min_row=2,
                                                             max_row=work_sheet.max_row,
                                                             min_col=2,
                                                             max_col=work_sheet.max_column)):
            name = att_name
            attendances = {}
            for date, cell in zip(attendance_dates, row):
                attendances[date] = cell.value
            attendance_data[name] = attendances
    else:
        print('ERROR: There is no attendance.xlsx file!!!')

    return attendance_data


def xlsx_file_write(path, data):
    xlsx_file_path = f'{path}/attendance.xlsx'
    work_book = xl.Workbook()
    work_sheet = work_book.active
    #  Write 1st row with dates
    rows = []
    row = ['NAMES']
    for item in data:
        for date in data[item]:
            row.append(date)
        rows.append(row)
        break

    # Fill other rows
    for item in data:
        row = [item]
        for date in data[item]:
            row.append(data[item][date])
        rows.append(row)
    # Put data into xlsx-file
    for row in rows:
        work_sheet.append(row)
    # Set styles to headers
    attendance_col_header_style(work_book, work_sheet)
    attendance_row_header_style(work_book, work_sheet)
    # Create a NamedStyle for cells (if not already defined)
    attendance_cell_style(work_book)
    # Create a NamedStyle for absent people cells (if not already defined)
    attendance_cell_style_absent(work_book)
    # Set styles to cells
    for row in work_sheet.iter_rows(min_row=2,
                                    max_row=work_sheet.max_row,
                                    min_col=2,
                                    max_col=work_sheet.max_column):
        for cell in row:
            if cell.value == 1:
                cell.style = 'attendance_cell'
            else:
                cell.style = 'attendance_cell_absent'
    # Adjust columns width according content
    xlsx_file_adjust_col_width(work_sheet)
    # Save xlsx-file
    work_book.save(xlsx_file_path)
    print(f'Data saved into file {xlsx_file_path}!')


def xlsx_file_create_new(path: str, class_names: List[str]) -> None:
    # Create XLSX-file
    def create_xlsx_file():
        # Create new xlsx-file
        work_book = xl.Workbook()
        work_sheet = work_book.active
        class_names.insert(0, 'NAMES')
        class_names.append('UNKNOWN')
        for name in class_names:
            work_sheet.append([name])
        #  Create a NamedStyle for cols (if not already defined)
        attendance_col_header_style(work_book, work_sheet)

        #  Create a NamedStyle for rows (if not already defined)
        attendance_row_header_style(work_book, work_sheet)
        xlsx_file_adjust_col_width(work_sheet)
        work_book.save(xlsx_file_path)
        print('XLSX-file created!')

    # Get a path to xlsx-file
    xlsx_file_path = f'{path}/attendance.xlsx'
    if not os.path.isfile(xlsx_file_path):
        # Create new xlsx-file
        create_xlsx_file()
    else:
        work_book = xl.load_workbook(xlsx_file_path)
        work_sheet = work_book.active
        if work_sheet['B1'].value:
            print('XLSX-file is not empty!')
        else:
            create_xlsx_file()


def xlsx_file_adjust_col_width(work_sheet):
    dims = {}
    for row in work_sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        work_sheet.column_dimensions[col].width = value * 1.2


def attendance_cell_style(work_book):
    if 'attendance_cell' not in work_book.named_styles:
        cell_style = NamedStyle(name='attendance_cell')
        cell_style.font = Font(bold=True, size=12)
        cell_style.alignment = Alignment(horizontal='center')
        cell_style.fill = PatternFill(fill_type='solid',
                                      start_color='00F0FDF0',
                                      end_color='00F0FDF0')
        cell_style.border = Border(
            left=Side(border_style='hair', color='FF000000'),
            right=Side(border_style='hair', color='FF000000'),
            top=Side(border_style='hair', color='FF000000'),
            bottom=Side(border_style='hair', color='FF000000'))
        work_book.add_named_style(cell_style)


def attendance_cell_style_absent(work_book):
    if 'attendance_cell_absent' not in work_book.named_styles:
        cell_style_absent = NamedStyle(name='attendance_cell_absent')
        cell_style_absent.font = Font(bold=True, size=12)
        cell_style_absent.alignment = Alignment(horizontal='center')
        cell_style_absent.fill = PatternFill(start_color='00FFDDDD',
                                             end_color='00FFDDDD',
                                             fill_type="solid")
        cell_style_absent.border = Border(
            left=Side(border_style='hair', color='FF000000'),
            right=Side(border_style='hair', color='FF000000'),
            top=Side(border_style='hair', color='FF000000'),
            bottom=Side(border_style='hair', color='FF000000'))
        work_book.add_named_style(cell_style_absent)


def attendance_col_header_style(work_book, work_sheet):
    if 'attendance_col_header' not in work_book.named_styles:
        col_header_style = NamedStyle(name='attendance_col_header')
        col_header_style.font = Font(bold=True, size=12)
        col_header_style.alignment = Alignment(horizontal='center')
        col_header_style.border = Border(
            left=Side(border_style='hair', color='FF000000'),
            right=Side(border_style='hair', color='FF000000'),
            top=Side(border_style='hair', color='FF000000'),
            bottom=Side(border_style='hair', color='FF000000'))
        work_book.add_named_style(col_header_style)
        # Set NamedStyle to 1st row
        for cell in work_sheet['1:1']:
            cell.style = 'attendance_col_header'


def attendance_row_header_style(work_book, work_sheet):
    if 'attendance_row_header' not in work_book.named_styles:
        row_header_style = NamedStyle(name='attendance_row_header')
        row_header_style.font = Font(bold=True, size=12)
        row_header_style.border = Border(
            left=Side(border_style='hair', color='FF000000'),
            right=Side(border_style='hair', color='FF000000'),
            top=Side(border_style='hair', color='FF000000'),
            bottom=Side(border_style='hair', color='FF000000'))
        work_book.add_named_style(row_header_style)
        # Set NamedStyle to 1st col
        for cell in work_sheet['A']:
            cell.style = 'attendance_row_header'


def xlsx_main(path, class_names):
    xlsx_file_path = f'{path}/attendance.xlsx'
    if os.path.isfile(xlsx_file_path):
        xlsx_file_read(xlsx_file_path)
    else:
        xlsx_file_create_new(xlsx_file_path, class_names)


if __name__ == '__main__':
    xlsx_main('zoom_beetroot')
